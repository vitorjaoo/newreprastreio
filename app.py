"""
app.py — Portal do Cliente (Flask)
Rotas:
  /               → login
  /dashboard      → cliente: cards de entrega
  /entrega/<id>   → cliente: detalhe da NF com rastreio, NF e boletos
  /financeiro     → cliente: títulos
  /trocar-senha   → cliente: troca de senha
  /admin          → admin: dashboard
  /admin/upload   → admin: upload de documentos
  /admin/rastreio → admin: gerenciar timeline de rastreio
  /admin/clientes → admin: clientes
  /admin/nfs      → admin: notas fiscais
  /admin/titulos  → admin: títulos
  /sair           → logout
"""

import base64
import hashlib
import io
import re
from functools import wraps
from datetime import datetime

from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, send_file, jsonify)

import db
from config import Config
from extrator_pdf import extrair_dados_nf, extrair_dados_boleto, pdf_para_base64

app = Flask(__name__)
app.config.from_object(Config)
app.secret_key = Config.SECRET_KEY

# Injeta variáveis globais em todos os templates
@app.context_processor
def globals_template():
    return {
        "nome_escritorio": Config.NOME_ESCRITORIO,
        "cor_primaria":    Config.COR_PRIMARIA,
        "cor_secundaria":  Config.COR_SECUNDARIA,
        "ano_atual":       datetime.now().year,
    }

# Inicializa banco na primeira requisição
with app.app_context():
    db.criar_tabelas()


# ─── Helpers ──────────────────────────────────────────────────────────────────

def hash_senha(s):
    return hashlib.sha256(s.encode()).hexdigest()

def login_cliente_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("perfil") != "cliente":
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

def login_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("perfil") != "admin":
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ═══════════════════════════════════════════════════════════════════════════════
# AUTH
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/", methods=["GET", "POST"])
def login():
    if session.get("perfil"):
        if session["perfil"] == "admin":
            return redirect(url_for("admin_dashboard"))
        return redirect(url_for("dashboard"))

    erro = None
    if request.method == "POST":
        cnpj  = request.form.get("cnpj", "").strip()
        senha = request.form.get("senha", "").strip()

        if cnpj.lower() == "admin" and senha == Config.ADMIN_SENHA:
            session["perfil"]  = "admin"
            session["usuario"] = {"nome": "Administrador", "id": 0}
            return redirect(url_for("admin_dashboard"))

        cliente = db.buscar_cliente_cnpj(cnpj)
        if not cliente:
            erro = "CNPJ não encontrado."
        elif hash_senha(senha) != cliente["senha_hash"]:
            erro = "Senha incorreta."
        elif not cliente["ativo"]:
            erro = "Acesso desativado. Entre em contato com o escritório."
        else:
            session["perfil"]  = "cliente"
            session["usuario"] = cliente
            return redirect(url_for("dashboard"))

    return render_template("login.html", erro=erro)


@app.route("/sair")
def sair():
    session.clear()
    return redirect(url_for("login"))


# ═══════════════════════════════════════════════════════════════════════════════
# CLIENTE
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/dashboard")
@login_cliente_required
def dashboard():
    cliente = session["usuario"]
    nfs     = db.listar_nfs(cliente["id"])
    titulos = db.listar_titulos(cliente["id"])

    hoje = datetime.now().strftime("%Y-%m-%d")
    titulos_abertos  = [t for t in titulos if t["status"] == "aberto"]
    titulos_vencidos = []
    for t in titulos_abertos:
        try:
            p = t["vencimento"].split("/")
            if f"{p[2]}-{p[1]}-{p[0]}" < hoje:
                titulos_vencidos.append(t)
        except Exception:
            pass

    # Adiciona eventos de rastreio em cada NF
    for nf in nfs:
        nf["eventos"] = db.listar_eventos_rastreio(nf["id"])
        nf["tem_rastreio"] = len(nf["eventos"]) > 0

    return render_template("dashboard.html",
        cliente=cliente,
        nfs=nfs,
        titulos_abertos=titulos_abertos,
        titulos_vencidos=titulos_vencidos,
    )


@app.route("/entrega/<int:nf_id>")
@login_cliente_required
def entrega(nf_id):
    cliente = session["usuario"]
    nfs     = db.listar_nfs(cliente["id"])
    nf      = next((n for n in nfs if n["id"] == nf_id), None)
    if not nf:
        flash("Nota fiscal não encontrada.", "erro")
        return redirect(url_for("dashboard"))

    nf["eventos"]    = db.listar_eventos_rastreio(nf_id)
    nf["pdf"]        = db.get_pdf_nf(nf_id)
    titulos_nf       = [t for t in db.listar_titulos(cliente["id"])
                        if t.get("nf_id") == nf_id]
    return render_template("entrega.html",
        cliente=cliente,
        nf=nf,
        titulos=titulos_nf,
    )


@app.route("/download/nf/<int:nf_id>")
@login_cliente_required
def download_nf(nf_id):
    cliente = session["usuario"]
    nfs     = db.listar_nfs(cliente["id"])
    if not any(n["id"] == nf_id for n in nfs):
        return "Não autorizado", 403
    dados = db.get_pdf_nf(nf_id)
    if not dados or not dados.get("pdf_base64"):
        return "PDF não disponível", 404
    pdf_bytes = base64.b64decode(dados["pdf_base64"])
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=dados.get("nome_arquivo") or f"NF_{nf_id}.pdf"
    )


@app.route("/download/boleto/<int:titulo_id>")
@login_cliente_required
def download_boleto(titulo_id):
    cliente = session["usuario"]
    titulos = db.listar_titulos(cliente["id"])
    if not any(t["id"] == titulo_id for t in titulos):
        return "Não autorizado", 403
    dados = db.get_pdf_titulo(titulo_id)
    if not dados or not dados.get("boleto_base64"):
        return "PDF não disponível", 404
    pdf_bytes = base64.b64decode(dados["boleto_base64"])
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=dados.get("nome_arquivo") or f"Boleto_{titulo_id}.pdf"
    )


@app.route("/financeiro")
@login_cliente_required
def financeiro():
    cliente = session["usuario"]
    titulos = db.listar_titulos(cliente["id"])
    hoje    = datetime.now().strftime("%Y-%m-%d")

    from datetime import date
    for t in titulos:
        try:
            p = t["vencimento"].split("/")
            iso = f"{p[2]}-{p[1]}-{p[0]}"
            venc_date = date.fromisoformat(iso)
            hoje_date = date.today()
            dias = (venc_date - hoje_date).days
            t["dias_vencimento"] = dias
            if t["status"] == "aberto" and iso < hoje:
                t["status_visual"] = "vencido"
            else:
                t["status_visual"] = t["status"]
        except Exception:
            t["status_visual"] = t["status"]
            t["dias_vencimento"] = None

    em_aberto = sum(float(t["valor"] or 0) for t in titulos if t["status"] == "aberto")
    quitado   = sum(float(t["valor"] or 0) for t in titulos if t["status"] == "pago")

    return render_template("financeiro.html",
        cliente=cliente,
        titulos=titulos,
        em_aberto=em_aberto,
        quitado=quitado,
    )


@app.route("/trocar-senha", methods=["GET", "POST"])
@login_cliente_required
def trocar_senha():
    cliente = session["usuario"]
    if request.method == "POST":
        atual  = request.form.get("senha_atual", "")
        nova   = request.form.get("senha_nova", "")
        conf   = request.form.get("confirmar", "")
        dados  = db.buscar_cliente_cnpj(cliente["cnpj"])
        if hash_senha(atual) != dados["senha_hash"]:
            flash("Senha atual incorreta.", "erro")
        elif nova != conf:
            flash("As senhas não coincidem.", "erro")
        elif len(nova) < 4:
            flash("A nova senha deve ter pelo menos 4 caracteres.", "erro")
        else:
            db.atualizar_senha(cliente["id"], hash_senha(nova))
            flash("Senha alterada com sucesso!", "sucesso")
            return redirect(url_for("dashboard"))
    return render_template("trocar_senha.html", cliente=cliente)


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin")
@login_admin_required
def admin_dashboard():
    nfs     = db.listar_todas_nfs()
    titulos = db.listar_todos_titulos()
    clientes = db.listar_clientes()
    em_aberto = sum(float(t["valor"] or 0) for t in titulos if t["status"] == "aberto")
    return render_template("admin/dashboard.html",
        total_clientes=len(clientes),
        total_nfs=len(nfs),
        titulos_abertos=len([t for t in titulos if t["status"] == "aberto"]),
        em_aberto=em_aberto,
    )


@app.route("/admin/upload", methods=["GET", "POST"])
@login_admin_required
def admin_upload():
    clientes  = db.listar_clientes()
    opcoes    = {str(c["id"]): c for c in clientes}
    tipo_ativo = request.args.get("tipo", "nf")
    msg        = None

    if request.method == "POST":
        tipo       = request.form.get("tipo")
        cliente_id = int(request.form.get("cliente_id"))
        arquivo    = request.files.get("arquivo")

        if not arquivo or arquivo.filename == "":
            flash("Selecione um arquivo PDF.", "erro")
            return redirect(url_for("admin_upload"))
        if not arquivo.filename.lower().endswith(".pdf"):
            flash(f"Arquivo inválido ({arquivo.filename}). Selecione um PDF.", "erro")
            return redirect(url_for("admin_upload"))

        pdf_bytes = arquivo.read()
        pdf_b64   = base64.b64encode(pdf_bytes).decode()

        if tipo == "nf":
            db.inserir_nf(
                cliente_id=cliente_id,
                numero_nf=request.form.get("numero_nf",""),
                valor=float(request.form.get("valor") or 0),
                data_emissao=request.form.get("data_emissao",""),
                pdf_base64=pdf_b64,
                nome_arquivo=arquivo.filename,
                status=request.form.get("status","ativo"),
                observacao=request.form.get("observacao",""),
                representada=request.form.get("representada",""),
            )
            flash(f"NF {request.form.get('numero_nf')} salva com sucesso!", "sucesso")

        elif tipo == "boleto":
            nf_id = request.form.get("nf_id")
            db.inserir_titulo(
                cliente_id=cliente_id,
                numero_titulo=request.form.get("numero_titulo",""),
                valor=float(request.form.get("valor") or 0),
                vencimento=request.form.get("vencimento",""),
                boleto_base64=pdf_b64,
                nome_arquivo=arquivo.filename,
                nf_id=int(nf_id) if nf_id else None,
            )
            flash(f"Boleto {request.form.get('numero_titulo')} salvo!", "sucesso")

        return redirect(url_for("admin_upload"))

    return render_template("admin/upload.html", clientes=clientes, tipo_ativo=tipo_ativo)


@app.route("/admin/nfs-do-cliente/<int:cliente_id>")
@login_admin_required
def admin_nfs_cliente(cliente_id):
    nfs = db.listar_nfs(cliente_id)
    return jsonify([{"id": n["id"], "numero_nf": n["numero_nf"],
                     "data_emissao": n["data_emissao"]} for n in nfs])


@app.route("/admin/rastreio")
@login_admin_required
def admin_rastreio():
    nfs = db.listar_todas_nfs()
    for nf in nfs:
        nf["eventos"] = db.listar_eventos_rastreio(nf["id"])
    clientes = sorted(list({n["cliente"] for n in nfs}))
    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    return render_template("admin/rastreio.html", nfs=nfs, clientes=clientes, now_str=now_str)


@app.route("/admin/rastreio/adicionar", methods=["POST"])
@login_admin_required
def admin_rastreio_adicionar():
    nf_id     = int(request.form.get("nf_id"))
    descricao = request.form.get("descricao","").strip()
    data_hora = request.form.get("data_hora","").strip()
    if descricao:
        db.inserir_evento_rastreio(nf_id, descricao, data_hora)
        flash("Evento adicionado!", "sucesso")
    return redirect(url_for("admin_rastreio") + f"#nf-{nf_id}")


@app.route("/admin/rastreio/remover/<int:evento_id>", methods=["POST"])
@login_admin_required
def admin_rastreio_remover(evento_id):
    db.deletar_evento_rastreio(evento_id)
    return redirect(request.referrer or url_for("admin_rastreio"))


@app.route("/admin/clientes", methods=["GET", "POST"])
@login_admin_required
def admin_clientes():
    if request.method == "POST":
        acao = request.form.get("acao")

        if acao == "cadastrar":
            try:
                db.criar_cliente(
                    request.form["nome"], request.form["cnpj"],
                    request.form.get("email",""), request.form.get("whatsapp",""),
                    hash_senha(request.form["senha"])
                )
                flash(f"Cliente {request.form['nome']} cadastrado!", "sucesso")
            except Exception as e:
                flash(f"Erro: CNPJ já cadastrado?", "erro")

        elif acao == "importar":
            arquivo = request.files.get("planilha")
            if arquivo:
                resultado = importar_clientes_excel(arquivo.read())
                if resultado["sucesso"]:
                    importados = duplicados = 0
                    for c in resultado["clientes"]:
                        try:
                            db.criar_cliente(c["nome"], c["cnpj"], c["email"],
                                             c["whatsapp"], hash_senha(c["senha"]))
                            importados += 1
                        except Exception:
                            duplicados += 1
                    msg = f"{importados} cliente(s) importado(s)."
                    if duplicados:
                        msg += f" {duplicados} já existiam."
                    flash(msg, "sucesso")
                else:
                    flash(f"Erro na planilha: {resultado.get('erro')}", "erro")

        elif acao == "inativar":
            db.toggle_cliente_ativo(int(request.form["cliente_id"]))
            flash("Status atualizado.", "sucesso")

        return redirect(url_for("admin_clientes"))

    clientes = db.listar_clientes()
    return render_template("admin/clientes.html", clientes=clientes)


@app.route("/admin/nfs", methods=["GET", "POST"])
@login_admin_required
def admin_nfs():
    if request.method == "POST":
        nf_id    = int(request.form["nf_id"])
        db.atualizar_status_nf(nf_id,
            request.form.get("status",""),
            request.form.get("observacao",""))
        db.atualizar_representada(nf_id, request.form.get("representada",""))
        flash("NF atualizada!", "sucesso")
        return redirect(url_for("admin_nfs"))

    nfs = db.listar_todas_nfs()
    return render_template("admin/nfs.html", nfs=nfs)


@app.route("/admin/titulos", methods=["GET", "POST"])
@login_admin_required
def admin_titulos():
    if request.method == "POST":
        db.marcar_titulo_pago(int(request.form["titulo_id"]))
        flash("Marcado como pago!", "sucesso")
        return redirect(url_for("admin_titulos"))

    titulos = db.listar_todos_titulos()
    em_aberto = sum(float(t["valor"] or 0) for t in titulos if t["status"] == "aberto")
    recebido  = sum(float(t["valor"] or 0) for t in titulos if t["status"] == "pago")
    return render_template("admin/titulos.html",
        titulos=titulos, em_aberto=em_aberto, recebido=recebido)


# ─── Helpers internos ─────────────────────────────────────────────────────────

def importar_clientes_excel(arquivo_bytes: bytes) -> dict:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(arquivo_bytes), data_only=True)
        ws = wb["Clientes"]
        clientes, erros = [], []
        for row in ws.iter_rows(min_row=6, values_only=True):
            nome, cnpj, email, whatsapp, senha, status = (list(row) + [None]*6)[:6]
            if not nome or not cnpj or not senha:
                continue
            cnpj_limpo = re.sub(r"\D", "", str(cnpj))
            if len(cnpj_limpo) != 14:
                erros.append(f"CNPJ inválido: {cnpj} ({nome})")
                continue
            cnpj_fmt = f"{cnpj_limpo[:2]}.{cnpj_limpo[2:5]}.{cnpj_limpo[5:8]}/{cnpj_limpo[8:12]}-{cnpj_limpo[12:]}"
            clientes.append({
                "nome": str(nome).strip(), "cnpj": cnpj_fmt,
                "email": str(email).strip() if email else "",
                "whatsapp": str(whatsapp).strip() if whatsapp else "",
                "senha": str(senha).strip(),
            })
        return {"clientes": clientes, "erros": erros, "sucesso": True}
    except Exception as e:
        return {"sucesso": False, "erro": str(e)}



@app.route("/admin/extrair-pdf", methods=["POST"])
@login_admin_required
def extrair_pdf():
    """Recebe PDF via AJAX, extrai dados com Gemini e retorna JSON"""
    arquivo = request.files.get("arquivo")
    tipo    = request.form.get("tipo", "nf")
    if not arquivo or arquivo.filename == "":
        return jsonify({"sucesso": False, "erro": "Nenhum arquivo"})
    pdf_bytes = arquivo.read()
    if tipo == "nf":
        dados = extrair_dados_nf(pdf_bytes, arquivo.filename)
    else:
        dados = extrair_dados_boleto(pdf_bytes, arquivo.filename)
    return jsonify(dados)


if __name__ == "__main__":
    app.run(debug=True, port=5000)
